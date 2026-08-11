"""Rotas de Clientes: listagem, cadastro, verificação de duplicidade, ficha,
edição, exclusão (exclusão só N1), exportação e IMPORTAÇÃO de .xlsx (v142 —
migração módulo a módulo rumo ao ERP unificado; importação só habilitada com
a env var MODO_MIGRACAO=true, pensada pra rodar só na instância unificada)."""
import io
import os
import random
from datetime import date, datetime
from flask import render_template, request, redirect, url_for, flash, jsonify, session, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from db import get_db, close_db
from config import CORES, CLIENTE, hoje_app
from auth import login_required, get_ctx, pode_excluir
from utils import audit_log


def migracao_habilitada():
    return os.environ.get('MODO_MIGRACAO') == 'true'


def _parse_data(val):
    """Aceita datetime/date (Excel converteu) ou string 'dd/mm/aaaa'. None se inválido."""
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()
    try:
        return datetime.strptime(str(val).strip(), '%d/%m/%Y').date().isoformat()
    except Exception:
        return None


def _parse_datahora(val):
    """Igual _parse_data mas devolve TIMESTAMP (usado no criado_em, preserva histórico)."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day).isoformat()
    s = str(val).strip()
    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).isoformat()
        except Exception:
            continue
    return None


@login_required
def clientes():
    conn = get_db(); cur = conn.cursor()
    hoje = hoje_app()
    # Período por data de CADASTRO do cliente — padrão: início do ano até hoje.
    data_inicio = request.args.get('data_inicio', hoje.strftime('%Y-01-01'))
    data_fim = request.args.get('data_fim', hoje.strftime('%Y-%m-%d'))
    try: date.fromisoformat(data_inicio)
    except ValueError: data_inicio = hoje.strftime('%Y-01-01')
    try: date.fromisoformat(data_fim)
    except ValueError: data_fim = hoje.strftime('%Y-%m-%d')
    cur.execute("SELECT * FROM clientes WHERE ativo=TRUE AND DATE(criado_em) BETWEEN %s AND %s ORDER BY nome", (data_inicio, data_fim))
    lista = [dict(c) for c in cur.fetchall()]
    cur.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(codigo FROM 2) AS INTEGER)), 0) as m FROM clientes WHERE codigo ~ '^C[0-9]+$'")
    n = cur.fetchone()['m']
    cur.close(); close_db(conn)
    for c in lista:
        p = c['nome'].split()
        c['iniciais'] = (p[0][0] + (p[1][0] if len(p) > 1 else p[0][-1])).upper()
    ctx = get_ctx(); ctx.update(clientes=lista, next_id=f"C{n+1}",
                                data_inicio=data_inicio, data_fim=data_fim,
                                modo_migracao=migracao_habilitada())
    return render_template('clientes.html', **ctx)


@login_required
def novo_cliente():
    nome = request.form.get('nome', '').strip()
    if not nome: flash('Nome obrigatorio.', 'erro'); return redirect(url_for('clientes'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id FROM clientes WHERE LOWER(TRIM(nome))=LOWER(TRIM(%s))", (nome,))
    if cur.fetchone():
        cur.close(); close_db(conn)
        flash('DUPLICADO_NOME||' + nome, 'erro'); return redirect(url_for('clientes'))
    cur.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(codigo FROM 2) AS INTEGER)), 0) as m FROM clientes WHERE codigo ~ '^C[0-9]+$'")
    n = cur.fetchone()['m']
    try:
        cur.execute("""INSERT INTO clientes (codigo,nome,cpf,data_nascimento,telefone,telefone2,
            cep,logradouro,numero,complemento,bairro,cidade,uf,promocoes,crediario,cor_avatar,usuario_id,usuario_nome)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (f"C{n+1}", nome,
             request.form.get('cpf', '').strip() or None,
             request.form.get('data_nascimento') or None,
             request.form.get('telefone', '').strip() or None,
             request.form.get('telefone2', '').strip() or None,
             request.form.get('cep', '').strip() or None,
             request.form.get('logradouro', '').strip() or None,
             request.form.get('numero', '').strip() or None,
             request.form.get('complemento', '').strip() or None,
             request.form.get('bairro', '').strip() or None,
             request.form.get('cidade', '').strip() or None,
             request.form.get('uf', '').strip() or None,
             request.form.get('promocoes', '0') == '1',
             request.form.get('crediario', '0') == '1',
             random.choice(CORES),
             session.get('uid'), session.get('nome')))
        conn.commit(); flash('SUCESSO||Cliente cadastrado!', 'ok')
    except Exception as e: conn.rollback(); flash(str(e), 'erro')
    finally: cur.close(); close_db(conn)
    return redirect(url_for('clientes'))


@login_required
def criar_cliente_rapido():
    """Cadastra um cliente e devolve JSON — usado para cadastrar SEM sair da venda.
    Se já existir um cliente com o mesmo nome, devolve o existente (não duplica)."""
    nome = request.form.get('nome', '').strip()
    if not nome:
        return jsonify({'ok': False, 'erro': 'Informe o nome do cliente.'})
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("SELECT id,codigo,nome,crediario FROM clientes WHERE LOWER(TRIM(nome))=LOWER(TRIM(%s))", (nome,))
        ex = cur.fetchone()
        if ex:
            c = dict(ex)
            return jsonify({'ok': True, 'existente': True,
                            'cliente': {'id': c['id'], 'codigo': c['codigo'], 'nome': c['nome'], 'crediario': bool(c['crediario'])}})
        cur.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(codigo FROM 2) AS INTEGER)), 0) as m FROM clientes WHERE codigo ~ '^C[0-9]+$'")
        n = cur.fetchone()['m']
        cur.execute("""INSERT INTO clientes (codigo,nome,cpf,data_nascimento,telefone,telefone2,
            cep,logradouro,numero,complemento,bairro,cidade,uf,promocoes,crediario,cor_avatar,usuario_id,usuario_nome)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id,codigo,nome,crediario""",
            (f"C{n+1}", nome,
             request.form.get('cpf', '').strip() or None,
             request.form.get('data_nascimento') or None,
             request.form.get('telefone', '').strip() or None,
             request.form.get('telefone2', '').strip() or None,
             request.form.get('cep', '').strip() or None,
             request.form.get('logradouro', '').strip() or None,
             request.form.get('numero', '').strip() or None,
             request.form.get('complemento', '').strip() or None,
             request.form.get('bairro', '').strip() or None,
             request.form.get('cidade', '').strip() or None,
             request.form.get('uf', '').strip() or None,
             request.form.get('promocoes', '0') == '1',
             request.form.get('crediario', '0') == '1',
             random.choice(CORES),
             session.get('uid'), session.get('nome')))
        row = dict(cur.fetchone())
        conn.commit()
        return jsonify({'ok': True, 'cliente': {'id': row['id'], 'codigo': row['codigo'],
                                                'nome': row['nome'], 'crediario': bool(row['crediario'])}})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'erro': str(e)})
    finally:
        cur.close(); close_db(conn)


@login_required
def verificar_cliente():
    campo = request.args.get('campo'); valor = request.args.get('valor', '').strip()
    if not campo or not valor: return jsonify({'ok': True})
    conn = get_db(); cur = conn.cursor()
    if campo == 'nome': cur.execute("SELECT id FROM clientes WHERE LOWER(nome)=LOWER(%s)", (valor,))
    elif campo == 'cpf': cur.execute("SELECT id FROM clientes WHERE cpf=%s", (valor,))
    elif campo == 'telefone': cur.execute("SELECT id FROM clientes WHERE telefone=%s OR telefone2=%s", (valor, valor))
    row = cur.fetchone(); cur.close(); close_db(conn)
    return jsonify({'ok': not bool(row)})


@login_required
def ficha_cliente(cid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM clientes WHERE id=%s", (cid,))
    row = cur.fetchone(); cur.close(); close_db(conn)
    if not row:
        flash('Cliente nao encontrado.', 'erro'); return redirect(url_for('clientes'))
    c = dict(row)
    p = c['nome'].split()
    c['iniciais'] = (p[0][0] + (p[1][0] if len(p) > 1 else p[0][-1])).upper()
    ctx = get_ctx(); ctx['c'] = c
    return render_template('ficha_cliente.html', **ctx)


@login_required
def editar_cliente(cid):
    conn = get_db(); cur = conn.cursor()
    if request.method == 'POST':
        cur.execute("""UPDATE clientes SET nome=%s,cpf=%s,data_nascimento=%s,telefone=%s,
            telefone2=%s,cep=%s,logradouro=%s,numero=%s,complemento=%s,bairro=%s,
            cidade=%s,uf=%s,promocoes=%s,crediario=%s WHERE id=%s""",
            (request.form.get('nome', '').strip(),
             request.form.get('cpf', '').strip() or None,
             request.form.get('data_nascimento') or None,
             request.form.get('telefone', '').strip() or None,
             request.form.get('telefone2', '').strip() or None,
             request.form.get('cep', '').strip() or None,
             request.form.get('logradouro', '').strip() or None,
             request.form.get('numero', '').strip() or None,
             request.form.get('complemento', '').strip() or None,
             request.form.get('bairro', '').strip() or None,
             request.form.get('cidade', '').strip() or None,
             request.form.get('uf', '').strip() or None,
             request.form.get('promocoes', '0') == '1',
             request.form.get('crediario', '0') == '1', cid))
        conn.commit(); cur.close(); close_db(conn)
        flash('Cliente atualizado!', 'ok')
        return redirect(url_for('ficha_cliente', cid=cid))
    cur.execute("SELECT * FROM clientes WHERE id=%s", (cid,))
    c = cur.fetchone(); cur.close(); close_db(conn)
    if not c:
        flash('Cliente nao encontrado.', 'erro'); return redirect(url_for('clientes'))
    ctx = get_ctx(); ctx['c'] = c
    return render_template('editar_cliente.html', **ctx)


@login_required
def excluir_cliente(cid):
    if not pode_excluir():
        flash('Apenas o Administrador N1 pode excluir dados.', 'erro'); return redirect(url_for('clientes'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT codigo,nome FROM clientes WHERE id=%s", (cid,))
    _old = cur.fetchone()
    cur.execute("DELETE FROM clientes WHERE id=%s", (cid,))
    audit_log(cur, 'EXCLUIR_CLIENTE', 'clientes', cid, dict(_old) if _old else None)
    conn.commit(); cur.close(); close_db(conn)
    flash('Cliente excluido.', 'ok')
    return redirect(url_for('clientes'))


@login_required
def exportar_clientes():
    """v142: gera um .xlsx com TODOS os clientes ativos (sem filtro de período —
    migração precisa do histórico completo) para download. Inclui o ID original e a
    loja de origem (identidade da instância), para rastreabilidade numa futura
    reimportação no ERP unificado."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM clientes WHERE ativo=TRUE ORDER BY nome")
    itens = [dict(c) for c in cur.fetchall()]
    cur.close(); close_db(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'

    headers = ['ID original', 'Loja de origem', 'Código', 'Nome', 'CPF', 'Data nascimento',
               'Telefone', 'Telefone 2', 'CEP', 'Logradouro', 'Número', 'Complemento',
               'Bairro', 'Cidade', 'UF', 'Aceita promoções', 'Crediário habilitado',
               'Cadastrado por', 'Data de cadastro']
    ws.append(headers)
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')

    loja_origem = CLIENTE.get('sigla') or CLIENTE.get('nome') or ''
    for item in itens:
        ws.append([
            item.get('id'), loja_origem, item.get('codigo'), item.get('nome') or '',
            item.get('cpf') or '',
            item['data_nascimento'].strftime('%d/%m/%Y') if item.get('data_nascimento') else '',
            item.get('telefone') or '', item.get('telefone2') or '', item.get('cep') or '',
            item.get('logradouro') or '', item.get('numero') or '', item.get('complemento') or '',
            item.get('bairro') or '', item.get('cidade') or '', item.get('uf') or '',
            'Sim' if item.get('promocoes') else 'Não',
            'Sim' if item.get('crediario') else 'Não',
            item.get('usuario_nome') or '',
            item['criado_em'].strftime('%d/%m/%Y %H:%M') if item.get('criado_em') else '',
        ])

    widths = [10, 14, 8, 26, 15, 13, 15, 15, 10, 26, 8, 16, 16, 16, 5, 12, 16, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    prefixo = (CLIENTE.get('sigla') or 'clientes').replace(' ', '_').replace('·', '').strip('_')
    nome_arquivo = f"clientes_{prefixo}_{hoje_app().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome_arquivo,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@login_required
def importar_clientes():
    """v142: importa um .xlsx exportado (e já tratado/deduplicado pelo usuário) pra dentro
    desta instância — pensada pra rodar SÓ no ERP unificado. Protegida por dois cadeados:
    a env var MODO_MIGRACAO=true (só existe no serviço unificado) e admin N1.

    Sem conceito de loja de origem — o ERP unificado trata todo cliente igual; a
    deduplicação (mesma pessoa nas duas lojas) já vem pronta na planilha. Idempotente:
    reimportar o mesmo arquivo não duplica (índice único em origem_id — a linha da
    planilha já importada antes é pulada)."""
    if not migracao_habilitada():
        flash('Importação não habilitada nesta instância.', 'erro'); return redirect(url_for('clientes'))
    if not pode_excluir():
        flash('Apenas o Administrador N1 pode importar dados.', 'erro'); return redirect(url_for('clientes'))
    conn = get_db(); cur = conn.cursor()
    if request.method == 'GET':
        cur.close(); close_db(conn)
        return render_template('importar_clientes.html', **get_ctx())

    try:
        arquivo = request.files.get('arquivo')
        if not arquivo or not arquivo.filename:
            flash('Selecione o arquivo .xlsx exportado.', 'erro'); return redirect(url_for('importar_clientes'))

        wb = load_workbook(io.BytesIO(arquivo.read()), data_only=True)
        ws = wb['Clientes'] if 'Clientes' in wb.sheetnames else wb.active
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            flash('Planilha vazia.', 'erro'); return redirect(url_for('importar_clientes'))
        cabecalho = [str(h).strip() if h is not None else '' for h in linhas[0]]
        col = {nome: i for i, nome in enumerate(cabecalho)}
        if 'Nome' not in col:
            flash('A planilha precisa ter a coluna "Nome" (use o arquivo gerado pelo botão Exportar).', 'erro')
            return redirect(url_for('importar_clientes'))

        def val(row, chave):
            i = col.get(chave)
            if i is None or i >= len(row):
                return None
            v = row[i]
            return v.strip() if isinstance(v, str) else v

        # Código NOVO (renumerado) pra este banco — o código original fica só em
        # origem_codigo, como referência.
        cur.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(codigo FROM 2) AS INTEGER)), 0) as m FROM clientes WHERE codigo ~ '^C[0-9]+$'")
        proximo_codigo = cur.fetchone()['m'] + 1

        importados = pulados = ignorados = 0
        for row in linhas[1:]:
            nome = val(row, 'Nome')
            if not nome or not str(nome).strip():
                ignorados += 1
                continue
            origem_id_raw = val(row, 'ID original')
            origem_id = int(origem_id_raw) if isinstance(origem_id_raw, (int, float)) else \
                        (int(origem_id_raw) if str(origem_id_raw or '').strip().isdigit() else None)
            promocoes = str(val(row, 'Aceita promoções') or '').strip().lower() == 'sim'
            crediario = str(val(row, 'Crediário habilitado') or '').strip().lower() == 'sim'
            criado_em = _parse_datahora(val(row, 'Data de cadastro'))
            novo_codigo = f"C{proximo_codigo}"
            proximo_codigo += 1
            cur.execute("""INSERT INTO clientes
                (nome,codigo,cpf,data_nascimento,telefone,telefone2,cep,logradouro,numero,
                 complemento,bairro,cidade,uf,promocoes,crediario,cor_avatar,usuario_id,usuario_nome,
                 origem_codigo,origem_id,criado_em)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,COALESCE(%s,CURRENT_TIMESTAMP))
                ON CONFLICT (origem_id) WHERE origem_id IS NOT NULL DO NOTHING
                RETURNING id""",
                (str(nome).strip(), novo_codigo, val(row, 'CPF'), _parse_data(val(row, 'Data nascimento')),
                 val(row, 'Telefone'), val(row, 'Telefone 2'), val(row, 'CEP'), val(row, 'Logradouro'),
                 val(row, 'Número'), val(row, 'Complemento'), val(row, 'Bairro'), val(row, 'Cidade'),
                 val(row, 'UF'), promocoes, crediario, random.choice(CORES), None,
                 val(row, 'Cadastrado por'), val(row, 'Código'), origem_id, criado_em))
            if cur.fetchone():
                importados += 1
            else:
                pulados += 1
        conn.commit()
        audit_log(cur, 'IMPORTAR_CLIENTES', 'clientes', None,
                  {'importados': importados, 'pulados': pulados})
        conn.commit()
        flash(f"Importação concluída: {importados} cliente(s) novo(s) importado(s), "
              f"{pulados} já existiam (pulado(s) automaticamente), "
              f"{ignorados} linha(s) vazia(s) ignorada(s).", 'ok')
    except Exception as e:
        conn.rollback(); flash(f'Erro na importação: {e}', 'erro')
    finally:
        cur.close(); close_db(conn)
    return redirect(url_for('importar_clientes'))


def register(app):
    app.add_url_rule('/clientes', 'clientes', clientes)
    app.add_url_rule('/clientes/exportar', 'exportar_clientes', exportar_clientes)
    app.add_url_rule('/clientes/importar', 'importar_clientes', importar_clientes, methods=['GET', 'POST'])
    app.add_url_rule('/clientes/novo', 'novo_cliente', novo_cliente, methods=['POST'])
    app.add_url_rule('/clientes/novo-rapido', 'criar_cliente_rapido', criar_cliente_rapido, methods=['POST'])
    app.add_url_rule('/clientes/verificar', 'verificar_cliente', verificar_cliente)
    app.add_url_rule('/clientes/<int:cid>', 'ficha_cliente', ficha_cliente)
    app.add_url_rule('/clientes/<int:cid>/editar', 'editar_cliente', editar_cliente, methods=['GET', 'POST'])
    app.add_url_rule('/clientes/<int:cid>/excluir', 'excluir_cliente', excluir_cliente, methods=['POST'])
