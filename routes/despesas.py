"""Rotas de Despesas: listagem com período e gráficos (mensal×avulsa, categorias,
formas), novo lançamento passo a passo (parcelado ou conta a pagar única),
pagamento de parcela (lança saída no caixa) e exclusão (só N1)."""
from datetime import datetime, date, timedelta
import calendar
import io
import secrets
from flask import render_template, request, redirect, url_for, session, flash, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from db import get_db, close_db
from config import hoje_app, fim_mes_app
from auth import login_required, get_ctx, pode_excluir, is_admin
from utils import parse_brl


def _add_months_clamped(base_date, months):
    """Soma meses mantendo o dia original quando possível e ajustando para o último dia do mês quando necessário."""
    idx = (base_date.year * 12 + (base_date.month - 1)) + months
    y = idx // 12
    m = idx % 12 + 1
    last = calendar.monthrange(y, m)[1]
    return date(y, m, min(base_date.day, last))


def _parse_iso_date(value, default=None):
    try:
        return date.fromisoformat(value)
    except Exception:
        return default or hoje_app()


def _editar_datas_pagas(cur):
    """v141: corrige a DATA de vencimento e/ou a DATA DE PAGAMENTO de parcelas JÁ PAGAS
    (campos pgvenc_<id>/pgdata_<id> + lista pagos_ids). Mantém o CAIXA coerente ajustando
    o criado_em do lançamento daquela parcela. O valor não é alterado (já está no caixa)."""
    for pid_s in [x for x in request.form.get('pagos_ids', '').split(',') if x.strip().isdigit()]:
        pid = int(pid_s)
        cur.execute("SELECT pago, despesa_id FROM despesa_parcelas WHERE id=%s", (pid,))
        r = cur.fetchone()
        if not r or not r['pago']:
            continue
        pgv = request.form.get(f'pgvenc_{pid}', '').strip()
        pgd = request.form.get(f'pgdata_{pid}', '').strip()
        try: date.fromisoformat(pgv)
        except Exception: pgv = None
        try: date.fromisoformat(pgd)
        except Exception: pgd = None
        sets, args = [], []
        if pgv: sets.append("data_vencimento=%s"); args.append(pgv)
        if pgd: sets.append("data_pagamento=%s"); args.append(pgd)
        if sets:
            args.append(pid)
            cur.execute("UPDATE despesa_parcelas SET " + ",".join(sets) + " WHERE id=%s", tuple(args))
        # Mantém a despesa-mãe coerente (usada na listagem, sobretudo em recorrentes).
        if pgv:
            cur.execute("UPDATE despesas SET data_vencimento=%s WHERE id=%s", (pgv, r['despesa_id']))
        # Caixa coerente: a saída daquela parcela passa a ter a data de pagamento corrigida.
        if pgd:
            cur.execute("UPDATE caixa SET criado_em=%s WHERE parcela_id=%s", (pgd, pid))


def _add_novos_vencimentos(cur, did, grupo):
    """v140: cria os novos vencimentos informados no detalhamento (edição).
    Lê add_count + add_data_<k>/add_valor_<k>. Recorrente (grupo != None): cada novo
    vencimento vira uma nova conta mensal no MESMO grupo (linha em despesas + parcela),
    permitindo estender a série quando ela chega ao fim. Parcelada/única: novas parcelas
    na própria despesa. Só cria linhas com data válida e valor > 0."""
    try:
        add_count = int(request.form.get('add_count', '0') or 0)
    except ValueError:
        add_count = 0
    if add_count <= 0:
        return
    novos = []
    for k in range(1, add_count + 1):
        pd = request.form.get(f'add_data_{k}', '').strip()
        pv = parse_brl(request.form.get(f'add_valor_{k}', '0'))
        try: date.fromisoformat(pd)
        except Exception: pd = None
        if pd and pv > 0:
            novos.append((pd, pv))
    if not novos:
        return
    if grupo:
        cur.execute("SELECT * FROM despesas WHERE recorrencia_grupo=%s ORDER BY id LIMIT 1", (grupo,))
        tpl_row = cur.fetchone()
        tpl = dict(tpl_row) if tpl_row else {}
        categoria = tpl.get('categoria') or request.form.get('categoria', '').strip() or None
        descricao = tpl.get('descricao') or request.form.get('descricao', '').strip() or None
        tipo = tpl.get('tipo') or 'mensal'
        obs = tpl.get('obs_retirada')
        rec_total = tpl.get('recorrencia_total')
        rec_base = tpl.get('recorrencia_base')
        cur.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(codigo FROM 2) AS INTEGER)),0) m FROM despesas WHERE codigo ~ '^D[0-9]+$'")
        nextn = cur.fetchone()['m'] + 1
        cur.execute("SELECT COALESCE(MAX(recorrencia_seq),0) mx FROM despesas WHERE recorrencia_grupo=%s", (grupo,))
        seq = cur.fetchone()['mx'] or 0
        for pd, pv in novos:
            seq += 1
            cur.execute("""INSERT INTO despesas
                (codigo,descricao,categoria,valor,data_despesa,forma_pagamento,tipo,
                 parcelado,num_parcelas,local_retirada,obs_retirada,status,data_vencimento,usuario_id,usuario_nome,
                 recorrente,recorrencia_grupo,recorrencia_seq,recorrencia_total,recorrencia_base)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (f"D{nextn}", descricao, categoria, pv, hoje_app().isoformat(), None, tipo,
                 False, 1, None, obs, 'pendente', pd,
                 session['uid'], session['nome'], True, grupo, seq, rec_total, rec_base))
            new_id = cur.fetchone()['id']
            cur.execute("INSERT INTO despesa_parcelas (despesa_id,numero,valor,data_vencimento) VALUES (%s,1,%s,%s)",
                        (new_id, pv, pd))
            nextn += 1
    else:
        cur.execute("SELECT COALESCE(MAX(numero),0) mx FROM despesa_parcelas WHERE despesa_id=%s", (did,))
        base_num = cur.fetchone()['mx']
        for i, (pd, pv) in enumerate(novos, start=1):
            cur.execute("INSERT INTO despesa_parcelas (despesa_id,numero,valor,data_vencimento) VALUES (%s,%s,%s,%s)",
                        (did, base_num + i, pv, pd))


@login_required
def despesas():
    conn = get_db(); cur = conn.cursor()
    hoje = hoje_app()
    data_inicio = request.args.get('data_inicio', hoje.strftime('%Y-01-01'))
    data_fim    = request.args.get('data_fim',    hoje.strftime('%Y-%m-%d'))
    try: date.fromisoformat(data_inicio)
    except: data_inicio = hoje.strftime('%Y-01-01')
    try: date.fromisoformat(data_fim)
    except: data_fim = hoje.strftime('%Y-%m-%d')
    # v143: a tabela principal mostra UMA LINHA POR PARCELA (não por despesa) — cada
    # conta a pagar/paga aparece no seu próprio vencimento/valor, sem agregação. Segue o
    # mesmo período do filtro De/Até: pendente entra pelo vencimento, paga pela data do
    # pagamento — igual ao critério dos cards de Total a pagar/Total pago. Ordena sempre
    # do vencimento mais próximo para o mais distante.
    cur.execute("""SELECT p.id AS parcela_id, p.numero, p.valor, p.data_vencimento, p.referencia, p.pago,
                          d.id AS despesa_id, d.codigo, d.categoria, d.descricao, d.tipo,
                          d.parcelado, d.num_parcelas
                   FROM despesa_parcelas p
                   JOIN despesas d ON d.id = p.despesa_id
                   WHERE (p.pago = FALSE AND DATE(p.data_vencimento) BETWEEN %s AND %s)
                      OR (p.pago = TRUE AND DATE(p.data_pagamento) BETWEEN %s AND %s)
                   ORDER BY p.data_vencimento ASC, d.id ASC, p.numero ASC""",
                (data_inicio, data_fim, data_inicio, data_fim))
    lista = [dict(r) for r in cur.fetchall()]
    def _norm_tipo(t):
        # v143: 'mensal' é o nome atual do tipo; 'fixa'/'fixo' ficam como sinônimo legado.
        return 'mensal' if (t or '').strip().lower() in ('mensal', 'fixa', 'fixo') else 'avulsa'
    for r in lista:
        r['tipo'] = _norm_tipo(r.get('tipo'))
        # REF: mês de competência (mensal) OU número da parcela/total (avulsa parcelada).
        if r.get('referencia'):
            r['ref_display'] = r['referencia'].strftime('%m/%Y')
        elif r.get('parcelado') and (r.get('num_parcelas') or 1) > 1:
            r['ref_display'] = f"{r['numero']:02d}/{r['num_parcelas']}"
        else:
            r['ref_display'] = None
    total = round(sum(float(r['valor'] or 0) for r in lista), 2)
    cur.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(codigo FROM 2) AS INTEGER)), 0) as m FROM despesas WHERE codigo ~ '^D[0-9]+$'")
    n = cur.fetchone()['m']
    # Categorias para o cadastro (ordenadas)
    cur.execute("SELECT nome FROM despesa_categorias WHERE ativo=TRUE ORDER BY nome")
    categorias = [r['nome'] for r in cur.fetchall()]
    # Contas a pagar (parcelas pendentes do período filtrado)
    cur.execute("""SELECT p.id as parcela_id, p.numero, p.valor, p.data_vencimento, p.referencia,
                          d.id as despesa_id, d.codigo, d.descricao, d.categoria,
                          d.forma_pagamento, d.local_retirada, d.num_parcelas
                   FROM despesa_parcelas p JOIN despesas d ON d.id=p.despesa_id
                   WHERE p.pago=FALSE
                     AND DATE(p.data_vencimento) BETWEEN %s AND %s
                   ORDER BY p.data_vencimento, d.id, p.numero""", (data_inicio, data_fim))
    a_pagar = [dict(r) for r in cur.fetchall()]
    for p in a_pagar:
        p['atrasada'] = bool(p['data_vencimento'] and p['data_vencimento'] < hoje)
    total_a_pagar = round(sum(float(p['valor'] or 0) for p in a_pagar), 2)
    n_atrasadas = sum(1 for p in a_pagar if p['atrasada'])

    # v116: contas pagas do período ao lado das pendentes.
    # Base: data_pagamento dentro do período filtrado, para o fechamento financeiro real.
    cur.execute("""SELECT p.id as parcela_id, p.numero, p.valor, p.data_vencimento, p.data_pagamento, p.referencia,
                          p.forma_pagamento as forma_pagamento_parcela, p.obs_pagamento,
                          d.id as despesa_id, d.codigo, d.descricao, d.categoria,
                          d.forma_pagamento, d.local_retirada, d.num_parcelas
                   FROM despesa_parcelas p JOIN despesas d ON d.id=p.despesa_id
                   WHERE p.pago=TRUE
                     AND DATE(p.data_pagamento) BETWEEN %s AND %s
                   ORDER BY p.data_pagamento DESC, d.id DESC, p.numero DESC""", (data_inicio, data_fim))
    pagas_periodo = [dict(r) for r in cur.fetchall()]
    total_pagas_periodo = round(sum(float(p['valor'] or 0) for p in pagas_periodo), 2)
    cur.close(); close_db(conn)
    ctx = get_ctx()
    ctx.update(lista=lista, total=total, data_inicio=data_inicio, data_fim=data_fim, next_cod=f"D{n+1}",
               categorias=categorias, a_pagar=a_pagar, total_a_pagar=total_a_pagar,
               pagas_periodo=pagas_periodo, total_pagas_periodo=total_pagas_periodo,
               n_a_pagar=len(a_pagar), n_atrasadas=n_atrasadas)
    return render_template('despesas.html', **ctx)


@login_required
def exportar_despesas():
    """Gera um .xlsx com as despesas do período escolhido (mesmo critério da tabela
    principal: pendente entra pelo vencimento, paga pela data de pagamento)."""
    conn = get_db(); cur = conn.cursor()
    hoje = hoje_app()
    data_inicio = request.args.get('data_inicio', hoje.strftime('%Y-01-01'))
    data_fim = request.args.get('data_fim', hoje.strftime('%Y-%m-%d'))
    try: date.fromisoformat(data_inicio)
    except ValueError: data_inicio = hoje.strftime('%Y-01-01')
    try: date.fromisoformat(data_fim)
    except ValueError: data_fim = hoje.strftime('%Y-%m-%d')
    cur.execute("""SELECT p.numero, p.valor, p.data_vencimento, p.referencia, p.pago,
                          p.data_pagamento, p.forma_pagamento as forma_pagamento_parcela,
                          d.codigo, d.categoria, d.descricao, d.tipo, d.parcelado, d.num_parcelas
                   FROM despesa_parcelas p
                   JOIN despesas d ON d.id = p.despesa_id
                   WHERE (p.pago = FALSE AND DATE(p.data_vencimento) BETWEEN %s AND %s)
                      OR (p.pago = TRUE AND DATE(p.data_pagamento) BETWEEN %s AND %s)
                   ORDER BY p.data_vencimento ASC, d.id ASC, p.numero ASC""",
                (data_inicio, data_fim, data_inicio, data_fim))
    lista = [dict(r) for r in cur.fetchall()]
    cur.close(); close_db(conn)

    def _norm_tipo(t):
        return 'Mensal' if (t or '').strip().lower() in ('mensal', 'fixa', 'fixo') else 'Avulsa'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Despesas'
    headers = ['Código', 'Tipo', 'Categoria', 'Descrição', 'Vencimento', 'REF', 'Valor (R$)',
               'Situação', 'Data de pagamento', 'Forma de pagamento']
    ws.append(headers)
    header_fill = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')

    for r in lista:
        if r.get('referencia'):
            ref = r['referencia'].strftime('%m/%Y')
        elif r.get('parcelado') and (r.get('num_parcelas') or 1) > 1:
            ref = f"{r['numero']:02d}/{r['num_parcelas']}"
        else:
            ref = ''
        ws.append([
            r.get('codigo'), _norm_tipo(r.get('tipo')), r.get('categoria') or '', r.get('descricao') or '',
            r['data_vencimento'].strftime('%d/%m/%Y') if r.get('data_vencimento') else '', ref,
            float(r.get('valor') or 0),
            'Pago' if r.get('pago') else 'A pagar',
            r['data_pagamento'].strftime('%d/%m/%Y') if r.get('data_pagamento') else '',
            (r.get('forma_pagamento_parcela') or '').replace('_', ' ').title(),
        ])
        ws.cell(row=ws.max_row, column=7).number_format = '#,##0.00'

    widths = [10, 9, 20, 28, 12, 10, 14, 10, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_arquivo = f"despesas_{data_inicio}_a_{data_fim}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome_arquivo,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@login_required
def nova_despesa():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(codigo FROM 2) AS INTEGER)), 0) as m FROM despesas WHERE codigo ~ '^D[0-9]+$'")
    n = cur.fetchone()['m']
    valor = parse_brl(request.form.get('valor', '0'))
    descricao = request.form.get('descricao', '').strip()
    categoria = request.form.get('categoria', '').strip()
    # v116: forma/origem/observação são informadas somente no pagamento.
    forma = ''
    tipo = request.form.get('tipo', 'avulsa').strip().lower()
    if tipo not in ('mensal', 'avulsa'): tipo = 'avulsa'
    local_ret = None
    obs_ret = None
    # v143: Mensal é sempre recorrente (gera as 12 contas mensais); parcelamento
    # só existe para Avulsa — o tipo já define o comportamento, sem pergunta extra.
    if tipo == 'mensal':
        recorrente = True
        parcelado = False
        num_parc = 1
        venc_unico = hoje_app().isoformat()
        venc_base = hoje_app()
    else:
        recorrente = False
        parcelado = request.form.get('parcelado', 'nao').strip().lower() == 'sim'
        try:
            num_parc = int(request.form.get('num_parcelas', '1') or 1)
        except ValueError:
            num_parc = 1
        if not parcelado or num_parc < 2:
            parcelado = False; num_parc = 1
        num_parc = min(max(num_parc, 1), 48)
        # Vencimento da conta a pagar única (avulsa sem parcelamento)
        venc_unico = request.form.get('data_vencimento_unica') or hoje_app().isoformat()
        venc_base = _parse_iso_date(venc_unico, hoje_app())
    # Rótulo da despesa para histórico (categoria + descrição livre)
    rotulo = categoria or descricao or 'Despesa'
    if categoria and descricao:
        rotulo = f"{categoria} — {descricao}"
    try:
        # Persistir categoria nova, se digitada
        if categoria:
            cur.execute("INSERT INTO despesa_categorias (nome) VALUES (%s) ON CONFLICT (nome) DO NOTHING", (categoria,))
        # Tanto parcelada quanto não-parcelada entram como conta(s) a pagar (pendente).
        # Recorrente: gera 12 lançamentos independentes, todos em aberto, permitindo ajustar valor/vencimento mês a mês.
        status = 'pendente'
        if recorrente:
            grupo = 'REC-' + hoje_app().strftime('%Y%m%d') + '-' + secrets.token_hex(4).upper()
            for idx in range(12):
                # v143: vencimento e REF (mês de competência) vêm da grade de 12 meses
                # preenchida no cadastro; cai para o cálculo automático (+1 mês) se faltar.
                venc_form = request.form.get(f'rec_data_{idx+1}', '').strip()
                venc = _parse_iso_date(venc_form, None) if venc_form else _add_months_clamped(venc_base, idx)
                ref_form = request.form.get(f'rec_ref_{idx+1}', '').strip()
                referencia = f"{ref_form}-01" if ref_form else None
                cod = f"D{n+1+idx}"
                cur.execute("""INSERT INTO despesas
                    (codigo,descricao,categoria,valor,data_despesa,forma_pagamento,tipo,
                     parcelado,num_parcelas,local_retirada,obs_retirada,status,data_vencimento,usuario_id,usuario_nome,
                     recorrente,recorrencia_grupo,recorrencia_seq,recorrencia_total,recorrencia_base)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (cod, descricao or None, categoria or None, valor,
                     request.form.get('data_despesa') or hoje_app().isoformat(),
                     forma or None, tipo, False, 1, local_ret, obs_ret, status, venc.isoformat(),
                     session['uid'], session['nome'], True, grupo, idx + 1, 12, venc_base.isoformat()))
                desp_id_mes = cur.fetchone()['id']
                cur.execute("""INSERT INTO despesa_parcelas (despesa_id,numero,valor,data_vencimento,referencia)
                    VALUES (%s,1,%s,%s,%s)""", (desp_id_mes, valor, venc.isoformat(), referencia))
            flash('Despesa recorrente registrada! Foram geradas 12 contas a pagar mensais em aberto.', 'ok')
        else:
            data_venc_desp = None if parcelado else venc_unico
            cur.execute("""INSERT INTO despesas
                (codigo,descricao,categoria,valor,data_despesa,forma_pagamento,tipo,
                 parcelado,num_parcelas,local_retirada,obs_retirada,status,data_vencimento,usuario_id,usuario_nome,
                 recorrente,recorrencia_total)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (f"D{n+1}", descricao or None, categoria or None, valor,
                 request.form.get('data_despesa') or hoje_app().isoformat(),
                 forma or None, tipo, parcelado, num_parc, local_ret, obs_ret, status, data_venc_desp,
                 session['uid'], session['nome'], False, 1))
            desp_id = cur.fetchone()['id']
        if (not recorrente) and parcelado:
            # Gera as parcelas; só viram saída no Caixa quando pagas
            soma = 0.0
            for i in range(1, num_parc + 1):
                pv = request.form.get(f'parcela_valor_{i}')
                pd = request.form.get(f'parcela_data_{i}')
                pvf = parse_brl(pv) if pv else round(valor / num_parc, 2)
                if not pd:
                    pd = (hoje_app() + timedelta(days=30 * i)).isoformat()
                soma += pvf
                cur.execute("""INSERT INTO despesa_parcelas (despesa_id,numero,valor,data_vencimento)
                    VALUES (%s,%s,%s,%s)""", (desp_id, i, pvf, pd))
            flash(f'Despesa parcelada em {num_parc}x registrada! As parcelas entram no caixa conforme você as paga.', 'ok')
        elif not recorrente:
            # Sem parcelamento = 1 conta a pagar com vencimento (só entra no caixa quando paga)
            cur.execute("""INSERT INTO despesa_parcelas (despesa_id,numero,valor,data_vencimento)
                VALUES (%s,1,%s,%s)""", (desp_id, valor, venc_unico))
            try: venc_fmt = datetime.fromisoformat(venc_unico).strftime('%d/%m/%Y')
            except Exception: venc_fmt = venc_unico
            flash(f'Despesa registrada como conta a pagar (vence em {venc_fmt}). Marque como paga quando quitar.', 'ok')
        conn.commit()
    except Exception as e: conn.rollback(); flash(str(e), 'erro')
    finally: cur.close(); close_db(conn)
    return redirect(url_for('despesas'))


@login_required
def pagar_parcela_despesa(did, pid):
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM despesas WHERE id=%s", (did,))
        d = cur.fetchone()
        if not d:
            flash('Despesa não encontrada.', 'erro'); return redirect(url_for('despesas'))
        d = dict(d)
        cur.execute("SELECT * FROM despesa_parcelas WHERE id=%s", (pid,))
        p = cur.fetchone()
        if not p:
            flash('Parcela não encontrada.', 'erro'); return redirect(url_for('despesas'))
        p = dict(p)
        if p['pago']:
            flash('Esta parcela já está paga.', 'erro'); return redirect(url_for('despesas'))
        data_pagamento = request.form.get('data_pagamento') or hoje_app().isoformat()
        forma_pagamento = (request.form.get('forma_pagamento') or '').strip()
        obs_pagamento = (request.form.get('obs_pagamento') or '').strip() or None
        if not forma_pagamento:
            flash('Informe a forma de pagamento para quitar a despesa.', 'erro')
            return redirect(url_for('despesas'))
        # valida data ISO; se vier inválida, usa a data atual do Brasil
        try:
            date.fromisoformat(data_pagamento)
        except Exception:
            data_pagamento = hoje_app().isoformat()
        cur.execute("""UPDATE despesa_parcelas
                       SET pago=TRUE,data_pagamento=%s,forma_pagamento=%s,obs_pagamento=%s
                       WHERE id=%s""", (data_pagamento, forma_pagamento, obs_pagamento, pid))
        rotulo = d.get('categoria') or d.get('descricao') or 'Despesa'
        descr_caixa = f"Despesa: {rotulo} (parc. {p['numero']}/{d.get('num_parcelas')})"
        if obs_pagamento:
            descr_caixa += f" — {obs_pagamento}"
        cur.execute("""INSERT INTO caixa
                       (descricao,valor,tipo,forma_pagamento,despesa_id,parcela_id,usuario_id,vendedora_nome,criado_em)
                       VALUES (%s,%s,'saida',%s,%s,%s,%s,%s,%s)""",
            (descr_caixa, float(p['valor']), forma_pagamento, did, pid, session['uid'], session['nome'], data_pagamento))
        # Se todas pagas, fecha a despesa
        cur.execute("SELECT COUNT(*) as t FROM despesa_parcelas WHERE despesa_id=%s AND pago=FALSE", (did,))
        if cur.fetchone()['t'] == 0:
            cur.execute("UPDATE despesas SET status='pago' WHERE id=%s", (did,))
        conn.commit(); flash(f"Parcela {p['numero']} paga e lançada no caixa!", 'ok')
    except Exception as e: conn.rollback(); flash(str(e), 'erro')
    finally: cur.close(); close_db(conn)
    return redirect(url_for('despesas'))


@login_required
def editar_despesa(did):
    """Edita uma despesa. Liberado para N1 e N2.
    Categoria/descrição/tipo/forma/origem podem ser alterados sempre.
    Valor e vencimento só quando a despesa ainda não foi paga e não é parcelada
    (conta a pagar única em aberto) — evita desencontro com parcelas/caixa."""
    if not is_admin():
        flash('Apenas administradores (N1 ou N2) podem editar despesas.', 'erro')
        return redirect(url_for('despesas'))
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM despesas WHERE id=%s", (did,))
        d = cur.fetchone()
        if not d:
            flash('Despesa não encontrada.', 'erro'); return redirect(url_for('despesas'))
        d = dict(d)
        categoria = request.form.get('categoria', '').strip()
        descricao = request.form.get('descricao', '').strip()
        forma = request.form.get('forma_pagamento', '').strip()
        tipo = request.form.get('tipo', 'avulsa').strip().lower()
        if tipo not in ('mensal', 'avulsa'): tipo = 'avulsa'
        local_ret = request.form.get('local_retirada', '').strip().lower()
        if local_ret not in ('caixa', 'pix'): local_ret = None
        obs_ret = request.form.get('obs_retirada', '').strip() or None
        # Valor/vencimento só são editáveis em conta a pagar única ainda em aberto
        editavel_valor = (not d.get('parcelado')) and (d.get('status') != 'pago')
        if categoria:
            cur.execute("INSERT INTO despesa_categorias (nome) VALUES (%s) ON CONFLICT (nome) DO NOTHING", (categoria,))
        if editavel_valor:
            valor = parse_brl(request.form.get('valor', '0'))
            venc = request.form.get('data_vencimento') or d.get('data_vencimento')
            cur.execute("""UPDATE despesas SET categoria=%s,descricao=%s,forma_pagamento=%s,tipo=%s,
                           local_retirada=%s,obs_retirada=%s,valor=%s,data_vencimento=%s WHERE id=%s""",
                        (categoria or None, descricao or None, forma or None, tipo, local_ret, obs_ret, valor, venc, did))
            # Atualiza a parcela única ainda em aberto (mantém a conta a pagar coerente)
            cur.execute("UPDATE despesa_parcelas SET valor=%s, data_vencimento=%s WHERE despesa_id=%s AND pago=FALSE",
                        (valor, venc, did))
        else:
            cur.execute("""UPDATE despesas SET categoria=%s,descricao=%s,forma_pagamento=%s,tipo=%s,
                           local_retirada=%s,obs_retirada=%s WHERE id=%s""",
                        (categoria or None, descricao or None, forma or None, tipo, local_ret, obs_ret, did))
        # Se a forma mudou, mantém o caixa coerente para lançamentos já pagos desta despesa
        cur.execute("UPDATE caixa SET forma_pagamento=%s WHERE despesa_id=%s", (forma or None, did))
        conn.commit(); flash('Despesa atualizada!', 'ok')
    except Exception as e:
        conn.rollback(); flash(str(e), 'erro')
    finally:
        cur.close(); close_db(conn)
    return redirect(url_for('despesas'))


@login_required
def detalhe_despesa(did):
    """JSON com tudo da despesa + parcelas — usado no detalhamento (clique na linha).

    v140: se a despesa for RECORRENTE (mensal que gerou 12 contas mensais no mesmo grupo),
    o detalhamento agrega TODAS as contas do grupo em uma única tela. Cada mês vira uma
    "parcela" editável (valor/vencimento), evitando ter de abrir mês a mês. Para despesas
    parceladas/únicas mantém o comportamento anterior (as parcelas da própria despesa)."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM despesas WHERE id=%s", (did,))
    d = cur.fetchone()
    if not d:
        cur.close(); close_db(conn); return jsonify({'ok': False})
    d = dict(d)
    def ds(x): return x.isoformat() if x else None
    grupo = d.get('recorrencia_grupo')
    out_parc = []
    if grupo:
        # Recorrente: junta as parcelas de TODAS as despesas do grupo, uma por mês.
        cur.execute("""SELECT p.*, p.despesa_id AS dp_id
                       FROM despesa_parcelas p
                       JOIN despesas dd ON dd.id = p.despesa_id
                       WHERE dd.recorrencia_grupo = %s
                       ORDER BY p.data_vencimento, p.despesa_id, p.numero""", (grupo,))
        parcelas = [dict(p) for p in cur.fetchall()]
        MESES_PT = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
        for i, p in enumerate(parcelas, start=1):
            v = p.get('data_vencimento')
            rot = f"{MESES_PT[v.month-1]}/{v.year}" if v else f"Mês {i}"
            out_parc.append({
                'id': p['id'], 'despesa_id': p['dp_id'], 'numero': i, 'rotulo_mes': rot,
                'valor': float(p['valor'] or 0), 'data_vencimento': ds(p.get('data_vencimento')),
                'referencia': ds(p.get('referencia')),
                'pago': bool(p['pago']), 'data_pagamento': ds(p.get('data_pagamento')),
                'forma_pagamento': p.get('forma_pagamento'), 'obs_pagamento': p.get('obs_pagamento'),
            })
    else:
        cur.execute("SELECT * FROM despesa_parcelas WHERE despesa_id=%s ORDER BY numero", (did,))
        parcelas = [dict(p) for p in cur.fetchall()]
        out_parc = [{
            'id': p['id'], 'despesa_id': did, 'numero': p['numero'], 'valor': float(p['valor'] or 0),
            'data_vencimento': ds(p.get('data_vencimento')), 'referencia': ds(p.get('referencia')),
            'pago': bool(p['pago']), 'data_pagamento': ds(p.get('data_pagamento')),
            'forma_pagamento': p.get('forma_pagamento'), 'obs_pagamento': p.get('obs_pagamento'),
        } for p in parcelas]
    total_geral = round(sum(p['valor'] for p in out_parc), 2) if grupo else float(d.get('valor') or 0)
    cur.close(); close_db(conn)
    return jsonify({'ok': True, 'despesa': {
        'id': d['id'], 'codigo': d.get('codigo'), 'categoria': d.get('categoria'),
        'descricao': d.get('descricao'), 'tipo': d.get('tipo'), 'valor': total_geral,
        'parcelado': bool(d.get('parcelado')),
        'num_parcelas': (len(out_parc) if grupo else d.get('num_parcelas')),
        'obs_retirada': d.get('obs_retirada'), 'data_despesa': ds(d.get('data_despesa')),
        'recorrente': bool(grupo), 'recorrencia_grupo': grupo,
    }, 'parcelas': out_parc})


@login_required
def salvar_despesa(did):
    """Salva a edição do detalhamento: cabeçalho (categoria/descrição/tipo/obs) e,
    principalmente, o VALOR e o VENCIMENTO das parcelas AINDA EM ABERTO (renegociação
    com o fornecedor). Parcelas já pagas não são alteradas (já estão no caixa)."""
    if not is_admin():
        flash('Apenas administradores (N1 ou N2) podem editar despesas.', 'erro')
        return redirect(url_for('despesas'))
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("SELECT id, recorrencia_grupo FROM despesas WHERE id=%s", (did,))
        row = cur.fetchone()
        if not row:
            flash('Despesa não encontrada.', 'erro'); return redirect(url_for('despesas'))
        grupo = dict(row).get('recorrencia_grupo')
        categoria = request.form.get('categoria', '').strip()
        descricao = request.form.get('descricao', '').strip()
        tipo = request.form.get('tipo', 'avulsa').strip().lower()
        if tipo not in ('mensal', 'avulsa'): tipo = 'avulsa'
        obs = request.form.get('obs_retirada', '').strip() or None
        if categoria:
            cur.execute("INSERT INTO despesa_categorias (nome) VALUES (%s) ON CONFLICT (nome) DO NOTHING", (categoria,))
        if grupo:
            # ── RECORRENTE: edição de TODOS os meses do grupo em uma tela só ──
            # Cabeçalho (categoria/descrição/tipo/obs) vale para todas as contas do grupo.
            cur.execute("""UPDATE despesas SET categoria=%s, descricao=%s, tipo=%s, obs_retirada=%s
                           WHERE recorrencia_grupo=%s""",
                        (categoria or None, descricao or None, tipo, obs, grupo))
            # Mapa das parcelas do grupo (id -> despesa_id) e quais estão em aberto.
            cur.execute("""SELECT p.id, p.despesa_id, p.pago
                           FROM despesa_parcelas p JOIN despesas dd ON dd.id=p.despesa_id
                           WHERE dd.recorrencia_grupo=%s""", (grupo,))
            grp_parc = {r['id']: {'did': r['despesa_id'], 'aberto': (not r['pago'])} for r in cur.fetchall()}
            # ── Exclusões: só meses EM ABERTO; os já pagos ficam (histórico/caixa) ──
            for pid_s in [x for x in request.form.get('del_ids', '').split(',') if x.strip().isdigit()]:
                pid = int(pid_s)
                info = grp_parc.get(pid)
                if not info or not info['aberto']:
                    continue
                cur.execute("DELETE FROM caixa WHERE parcela_id=%s", (pid,))
                cur.execute("DELETE FROM despesa_parcelas WHERE id=%s", (pid,))
                cur.execute("DELETE FROM despesas WHERE id=%s", (info['did'],))  # cada mês é 1 despesa
                grp_parc.pop(pid, None)
            # ── Edições de valor/vencimento dos meses em aberto ──
            for pid_s in [x for x in request.form.get('parcela_ids', '').split(',') if x.strip().isdigit()]:
                pid = int(pid_s)
                info = grp_parc.get(pid)
                if not info or not info['aberto']:   # inexistente, já paga ou excluída → não mexe
                    continue
                pv = parse_brl(request.form.get(f'pval_{pid}', '0'))
                pdta = request.form.get(f'pdata_{pid}', '').strip()
                try: date.fromisoformat(pdta)
                except Exception: pdta = None
                if pv > 0 and pdta:
                    cur.execute("UPDATE despesa_parcelas SET valor=%s, data_vencimento=%s WHERE id=%s", (pv, pdta, pid))
                    cur.execute("UPDATE despesas SET valor=%s, data_vencimento=%s WHERE id=%s", (pv, pdta, info['did']))
                elif pv > 0:
                    cur.execute("UPDATE despesa_parcelas SET valor=%s WHERE id=%s", (pv, pid))
                    cur.execute("UPDATE despesas SET valor=%s WHERE id=%s", (pv, info['did']))
                elif pdta:
                    cur.execute("UPDATE despesa_parcelas SET data_vencimento=%s WHERE id=%s", (pdta, pid))
                    cur.execute("UPDATE despesas SET data_vencimento=%s WHERE id=%s", (pdta, info['did']))
            # ── Adições: novos vencimentos (novas contas mensais dentro do mesmo grupo) ──
            _add_novos_vencimentos(cur, did, grupo)
            # ── Correção de datas de meses JÁ PAGOS (vencimento / data de pagamento) ──
            _editar_datas_pagas(cur)
            conn.commit(); flash('Despesa recorrente atualizada!', 'ok')
            return redirect(url_for('despesas'))
        reparc_n = request.form.get('reparc_n', '').strip()
        if reparc_n.isdigit() and int(reparc_n) >= 1:
            # RENEGOCIAÇÃO (editor v140): apaga as parcelas EM ABERTO e cria N novas com os
            # valores/vencimentos informados. As já pagas continuam (histórico/caixa intactos).
            # O total (valor) da despesa é recalculado da soma logo abaixo.
            n = min(int(reparc_n), 48)
            cur.execute("DELETE FROM despesa_parcelas WHERE despesa_id=%s AND pago=FALSE", (did,))
            cur.execute("SELECT COALESCE(MAX(numero),0) mx FROM despesa_parcelas WHERE despesa_id=%s", (did,))
            base_num = cur.fetchone()['mx']
            for i in range(1, n + 1):
                pv = parse_brl(request.form.get(f'np_valor_{i}', '0'))
                pd = request.form.get(f'np_data_{i}', '').strip()
                try: date.fromisoformat(pd)
                except Exception: pd = None
                cur.execute("INSERT INTO despesa_parcelas (despesa_id,numero,valor,data_vencimento) VALUES (%s,%s,%s,%s)",
                            (did, base_num + i, pv, pd))
            cur.execute("UPDATE despesas SET parcelado=TRUE WHERE id=%s", (did,))
        else:
            # Parcelas em aberto: atualiza valor/vencimento (as pagas ficam intactas).
            cur.execute("SELECT id, pago FROM despesa_parcelas WHERE despesa_id=%s", (did,))
            aberto = {r['id']: (not r['pago']) for r in cur.fetchall()}
            # ── Exclusões: só parcelas EM ABERTO desta despesa ──
            for pid_s in [x for x in request.form.get('del_ids', '').split(',') if x.strip().isdigit()]:
                pid = int(pid_s)
                if not aberto.get(pid):
                    continue
                cur.execute("DELETE FROM caixa WHERE parcela_id=%s", (pid,))
                cur.execute("DELETE FROM despesa_parcelas WHERE id=%s", (pid,))
                aberto.pop(pid, None)
            for pid_s in [x for x in request.form.get('parcela_ids', '').split(',') if x.strip().isdigit()]:
                pid = int(pid_s)
                if not aberto.get(pid):   # inexistente, já paga ou excluída → não mexe
                    continue
                pv = parse_brl(request.form.get(f'pval_{pid}', '0'))
                pdta = request.form.get(f'pdata_{pid}', '').strip()
                try: date.fromisoformat(pdta)
                except Exception: pdta = None
                if pv > 0 and pdta:
                    cur.execute("UPDATE despesa_parcelas SET valor=%s, data_vencimento=%s WHERE id=%s", (pv, pdta, pid))
                elif pv > 0:
                    cur.execute("UPDATE despesa_parcelas SET valor=%s WHERE id=%s", (pv, pid))
                elif pdta:
                    cur.execute("UPDATE despesa_parcelas SET data_vencimento=%s WHERE id=%s", (pdta, pid))
            # ── Adições: novos vencimentos como parcelas da MESMA despesa ──
            _add_novos_vencimentos(cur, did, None)
        # ── Correção de datas de parcelas JÁ PAGAS (vencimento / data de pagamento) ──
        # Vale para parceladas e para a conta única quitada (roda em reparc e no else).
        _editar_datas_pagas(cur)
        # Recalcula total, próximo vencimento e nº de parcelas da despesa
        cur.execute("""SELECT COALESCE(SUM(valor),0) tot, COUNT(*) cnt,
                              MIN(CASE WHEN pago=FALSE THEN data_vencimento END) prox
                       FROM despesa_parcelas WHERE despesa_id=%s""", (did,))
        agg = cur.fetchone()
        cnt = int(agg['cnt'] or 1)
        cur.execute("UPDATE despesas SET categoria=%s, descricao=%s, tipo=%s, obs_retirada=%s, valor=%s, num_parcelas=%s, parcelado=%s WHERE id=%s",
                    (categoria or None, descricao or None, tipo, obs, float(agg['tot'] or 0), cnt, cnt > 1, did))
        if agg['prox'] is not None:
            cur.execute("UPDATE despesas SET data_vencimento=%s WHERE id=%s", (agg['prox'], did))
        conn.commit(); flash('Despesa atualizada!', 'ok')
    except Exception as e:
        conn.rollback(); flash(str(e), 'erro')
    finally:
        cur.close(); close_db(conn)
    return redirect(url_for('despesas'))


@login_required
def excluir_despesa(did):
    if not pode_excluir():
        flash('Apenas o Administrador N1 pode excluir dados.', 'erro'); return redirect(url_for('despesas'))
    conn = get_db(); cur = conn.cursor()
    try:
        # v140: se for recorrente, exclui TODA a série (o detalhamento mostra o grupo inteiro).
        cur.execute("SELECT recorrencia_grupo FROM despesas WHERE id=%s", (did,))
        row = cur.fetchone()
        grupo = dict(row).get('recorrencia_grupo') if row else None
        if grupo:
            cur.execute("SELECT id FROM despesas WHERE recorrencia_grupo=%s", (grupo,))
            ids = [r['id'] for r in cur.fetchall()]
            if ids:
                cur.execute("DELETE FROM caixa WHERE despesa_id = ANY(%s)", (ids,))
                cur.execute("DELETE FROM despesa_parcelas WHERE despesa_id = ANY(%s)", (ids,))
                cur.execute("DELETE FROM despesas WHERE id = ANY(%s)", (ids,))
            conn.commit(); flash('Despesa recorrente excluída (todos os meses).', 'ok')
            return redirect(url_for('despesas'))
        cur.execute("DELETE FROM caixa WHERE despesa_id=%s", (did,))
        cur.execute("DELETE FROM despesa_parcelas WHERE despesa_id=%s", (did,))
        cur.execute("DELETE FROM despesas WHERE id=%s", (did,))
        conn.commit(); flash('Despesa excluida.', 'ok')
    except Exception as e: conn.rollback(); flash(str(e), 'erro')
    finally: cur.close(); close_db(conn)
    return redirect(url_for('despesas'))


def register(app):
    app.add_url_rule('/despesas', 'despesas', despesas)
    app.add_url_rule('/despesas/exportar', 'exportar_despesas', exportar_despesas)
    app.add_url_rule('/despesas/nova', 'nova_despesa', nova_despesa, methods=['POST'])
    app.add_url_rule('/despesas/<int:did>/parcela/<int:pid>/pagar', 'pagar_parcela_despesa', pagar_parcela_despesa, methods=['POST'])
    app.add_url_rule('/despesas/<int:did>/editar', 'editar_despesa', editar_despesa, methods=['POST'])
    app.add_url_rule('/despesas/<int:did>/detalhe', 'detalhe_despesa', detalhe_despesa)
    app.add_url_rule('/despesas/<int:did>/salvar', 'salvar_despesa', salvar_despesa, methods=['POST'])
    app.add_url_rule('/despesas/<int:did>/excluir', 'excluir_despesa', excluir_despesa, methods=['POST'])
