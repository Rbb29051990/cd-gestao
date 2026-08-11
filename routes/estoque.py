"""Rotas de Estoque: listagem, cadastro, nova entrada, modelos/tamanhos,
etiquetas (lista por data e busca por código), ficha, edição, exclusão,
exportação para .xlsx, extração das fotos por código e IMPORTAÇÃO pro ERP
unificado (v142/v143 — importação só habilitada com MODO_MIGRACAO=true)."""
import base64
import io
import os
import zipfile
from datetime import date, datetime
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from db import get_db, close_db
from config import hoje_app, fim_mes_app
from auth import login_required, get_ctx, pode_excluir
from utils import parse_brl, audit_log

# v142: tipo do produto escolhido no cadastro define o prefixo do código — cada
# prefixo tem sua própria sequência numérica (ACn/PSn/SLn podem coexistir).
PREFIXOS_TIPO = {'Acessórios': 'AC', 'Plus Size': 'PS', 'Slim': 'SL'}


def migracao_habilitada():
    return os.environ.get('MODO_MIGRACAO') == 'true'


def _parse_data_planilha(val):
    """Aceita datetime/date (Excel converteu) ou string 'dd/mm/aaaa'. None se inválido."""
    if not val:
        return None
    if isinstance(val, (datetime, date)):
        return val.date().isoformat() if isinstance(val, datetime) else val.isoformat()
    try:
        return datetime.strptime(str(val).strip(), '%d/%m/%Y').date().isoformat()
    except Exception:
        return None


def _proximo_codigo(cur, prefixo):
    """Próximo código pro prefixo dado (AC/PS/SL) — sequências independentes."""
    cur.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(codigo FROM %s) AS INTEGER)), 0) as m "
                "FROM estoque WHERE codigo ~ %s", (len(prefixo) + 1, f'^{prefixo}[0-9]+$'))
    return cur.fetchone()['m'] + 1


def _foto_bytes_ext(foto):
    """Decodifica o data URI (base64) de uma foto -> (bytes, extensão) ou None se
    não tiver foto/for inválida. Usado tanto na exportação de fotos por planilha
    quanto na exportação combinada dos itens marcados na tela."""
    if not foto or not foto.startswith('data:'):
        return None
    try:
        header, b64 = foto.split(',', 1)
        mime = header.split(';')[0].replace('data:', '') or 'image/jpeg'
        ext = mime.split('/')[-1].lower()
        if ext == 'jpeg': ext = 'jpg'
        return base64.b64decode(b64), ext
    except Exception:
        return None


def _montar_wb_estoque(itens, entradas_map, hoje):
    """Monta o workbook de exportação do Estoque (mesmas colunas do botão
    'Exportar dados') a partir de uma lista de itens já carregada — reaproveitado
    pela exportação por período E pela exportação dos itens marcados na tela."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Estoque'

    headers = ['Código', 'Categoria', 'Data lançamento', 'Cadastrado por', 'Modelo', 'Descrição', 'Tamanho',
               'Estoque inicial', 'Entradas adicionais', 'Saídas', 'Saldo atual',
               'Custo unitário (R$)', 'Markup (%)', 'Valor de venda (R$)',
               'Margem de lucro (%)', 'Desconto promo (%)', 'Valor promocional (R$)',
               'Dias em estoque', 'Custo total do saldo (R$)', 'Valor total do saldo (R$)']
    ws.append(headers)
    header_fill = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')

    money_cols = {12, 14, 17, 19, 20}
    pct_cols = {13, 15, 16}
    for item in itens:
        entradas_adicionais = entradas_map.get(item['id'], 0)
        saldo = int(item.get('quantidade') or 0)
        saidas = max(0, (item.get('estoque_inicial') or 0) + entradas_adicionais - saldo)
        dias = (hoje - item['criado_em'].date()).days if item.get('criado_em') else None
        custo = float(item.get('custo_unitario') or 0)
        venda = float(item.get('valor_venda') or 0)
        dp = float(item.get('desconto_promo') or 0)
        valor_promo = round(venda * (1 - dp / 100), 2) if dp > 0 else None
        ws.append([
            item.get('codigo'), item.get('tipo_produto') or '',
            item['criado_em'].strftime('%d/%m/%Y %H:%M:%S') if item.get('criado_em') else '',
            item.get('usuario_nome') or '—',
            item.get('modelo') or '', item.get('descricao') or '', item.get('tamanho') or '',
            item.get('estoque_inicial') or 0, entradas_adicionais, saidas, saldo,
            custo, float(item.get('markup') or 0), venda, float(item.get('margem_lucro') or 0),
            dp if dp > 0 else None, valor_promo, dias,
            round(custo * saldo, 2), round(venda * saldo, 2),
        ])
        row = ws.max_row
        for col in money_cols:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        for col in pct_cols:
            ws.cell(row=row, column=col).number_format = '0.00'

    widths = [10, 10, 18, 18, 20, 28, 9, 10, 12, 9, 10, 15, 10, 16, 14, 13, 16, 12, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return wb


@login_required
def estoque():
    conn = get_db(); cur = conn.cursor()
    hoje = hoje_app()
    # Filtro de período por DATA DE LANÇAMENTO do produto (criado_em) — padrão das
    # demais abas: começa no mês vigente (1º dia → último dia).
    data_inicio = request.args.get('data_inicio', hoje.strftime('%Y-01-01'))
    data_fim = request.args.get('data_fim', hoje.strftime('%Y-%m-%d'))
    try: date.fromisoformat(data_inicio)
    except ValueError: data_inicio = hoje.strftime('%Y-01-01')
    try: date.fromisoformat(data_fim)
    except ValueError: data_fim = hoje.strftime('%Y-%m-%d')
    cur.execute("SELECT * FROM estoque WHERE ativo=TRUE AND DATE(criado_em) BETWEEN %s AND %s ORDER BY criado_em", (data_inicio, data_fim))
    itens = [dict(i) for i in cur.fetchall()]
    cur.execute("SELECT COALESCE(SUM(custo_unitario*quantidade),0) as ct, COALESCE(SUM(valor_venda*quantidade),0) as vt FROM estoque WHERE ativo=TRUE AND DATE(criado_em) BETWEEN %s AND %s", (data_inicio, data_fim))
    tots = cur.fetchone()
    cur.execute("SELECT nome FROM modelos_estoque ORDER BY nome")
    modelos = [r['nome'] for r in cur.fetchall()]
    cur.execute("SELECT nome FROM tamanhos_estoque ORDER BY id")
    tamanhos = [r['nome'] for r in cur.fetchall()]
    # Preview do próximo código de cada tipo — o JS troca qual mostrar conforme o
    # botão escolhido no cadastro.
    next_refs = {tipo: f"{pfx}{_proximo_codigo(cur, pfx)}" for tipo, pfx in PREFIXOS_TIPO.items()}
    # Buscar total de entradas adicionais por item (mesma conexão, antes de fechar)
    cur.execute("SELECT estoque_id, COALESCE(SUM(quantidade),0) as total FROM estoque_entradas GROUP BY estoque_id")
    entradas_map = {r['estoque_id']: int(r['total']) for r in cur.fetchall()}
    cur.close(); close_db(conn)
    for i in itens:
        i['dias_estoque'] = (hoje - i['criado_em'].date()).days
        i['entradas_adicionais'] = entradas_map.get(i['id'], 0)
        i['saidas'] = max(0, (i['estoque_inicial'] or 0) + i['entradas_adicionais'] - i['quantidade'])
        dp = float(i.get('desconto_promo') or 0)
        i['valor_promo'] = round(float(i['valor_venda'] or 0) * (1 - dp / 100), 2) if dp > 0 else None
    ctx = get_ctx()
    ctx.update(itens=itens, modelos=modelos, tamanhos=tamanhos,
               custo_total=float(tots['ct']), valor_total=float(tots['vt']),
               lucro_potencial=float(tots['vt']) - float(tots['ct']),
               next_refs=next_refs, tipos_produto=list(PREFIXOS_TIPO.keys()),
               data_inicio=data_inicio, data_fim=data_fim,
               modo_migracao=migracao_habilitada())
    return render_template('estoque.html', **ctx)


@login_required
def novo_estoque():
    conn = get_db(); cur = conn.cursor()
    tipo_produto = request.form.get('tipo_produto', '').strip()
    prefixo = PREFIXOS_TIPO.get(tipo_produto)
    if not prefixo:
        flash('Selecione o tipo do produto (Acessórios, Plus Size ou Slim).', 'erro')
        cur.close(); close_db(conn)
        return redirect(url_for('estoque'))
    codigo = f"{prefixo}{_proximo_codigo(cur, prefixo)}"
    qtd = int(request.form.get('quantidade', 1) or 1)
    custo_raw = request.form.get('custo_unitario', '').strip()
    venda_raw = request.form.get('valor_venda', '').strip()
    if not custo_raw or parse_brl(custo_raw) <= 0:
        flash('O custo unitário é obrigatório e deve ser maior que zero.', 'erro')
        cur.close(); close_db(conn)
        return redirect(url_for('estoque'))
    if not venda_raw or parse_brl(venda_raw) <= 0:
        flash('O valor de venda é obrigatório e deve ser maior que zero.', 'erro')
        cur.close(); close_db(conn)
        return redirect(url_for('estoque'))
    foto = request.form.get('foto', '').strip() or None
    # Segurança: só aceita data URI de imagem e limita o tamanho (~1.5MB de base64)
    if foto and (not foto.startswith('data:image/') or len(foto) > 1_500_000):
        foto = None
    try:
        cur.execute("""INSERT INTO estoque (codigo,tipo_produto,modelo,descricao,tamanho,quantidade,estoque_inicial,
            custo_unitario,markup,valor_venda,margem_lucro,foto,usuario_id,usuario_nome) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (codigo, tipo_produto, request.form.get('modelo', '').strip(),
             request.form.get('descricao', '').strip() or None,
             request.form.get('tamanho', '').strip(), qtd, qtd,
             parse_brl(request.form.get('custo_unitario', '0')),
             parse_brl(request.form.get('markup', '0')),
             parse_brl(request.form.get('valor_venda', '0')),
             parse_brl(request.form.get('margem_lucro', '0')), foto,
             session.get('uid'), session.get('nome')))
        conn.commit(); flash('Produto cadastrado!', 'ok')
    except Exception as e: conn.rollback(); flash(str(e), 'erro')
    finally: cur.close(); close_db(conn)
    return redirect(url_for('estoque'))


@login_required
def nova_entrada_estoque(eid):
    conn = get_db(); cur = conn.cursor()
    try:
        qtd = int(request.form.get('quantidade', 0) or 0)
        custo = parse_brl(request.form.get('custo_unitario', '0'))
        markup = parse_brl(request.form.get('markup', '0'))
        venda = parse_brl(request.form.get('valor_venda', '0'))
        margem = parse_brl(request.form.get('margem_lucro', '0'))
        if qtd <= 0:
            flash('Informe uma quantidade válida.', 'erro')
            return redirect(url_for('ficha_estoque', eid=eid))
        if custo <= 0:
            flash('O custo unitário é obrigatório para nova entrada.', 'erro')
            return redirect(url_for('ficha_estoque', eid=eid))
        if venda <= 0:
            flash('O valor de venda é obrigatório para nova entrada.', 'erro')
            return redirect(url_for('ficha_estoque', eid=eid))
        # Registrar entrada
        cur.execute("""INSERT INTO estoque_entradas (estoque_id, quantidade, custo_unitario, valor_venda, markup, margem_lucro)
                       VALUES (%s,%s,%s,%s,%s,%s)""", (eid, qtd, custo, venda, markup, margem))
        # Atualizar saldo e último custo/venda do produto
        cur.execute("""UPDATE estoque SET
                       quantidade = quantidade + %s,
                       custo_unitario = %s,
                       valor_venda = %s,
                       markup = %s,
                       margem_lucro = %s
                       WHERE id = %s""", (qtd, custo, venda, markup, margem, eid))
        conn.commit()
        flash(f'Nova entrada de {qtd} unidade(s) registrada com sucesso!', 'ok')
    except Exception as e:
        conn.rollback(); flash(str(e), 'erro')
    finally: cur.close(); close_db(conn)
    return redirect(url_for('ficha_estoque', eid=eid))


@login_required
def novo_modelo():
    nome = request.form.get('nome', '').strip()
    if nome:
        conn = get_db(); cur = conn.cursor()
        cur.execute("INSERT INTO modelos_estoque (nome) VALUES (%s) ON CONFLICT DO NOTHING", (nome,))
        conn.commit(); cur.close(); close_db(conn)
    return redirect(url_for('estoque'))


@login_required
def novo_tamanho():
    nome = request.form.get('nome', '').strip()
    if not nome:
        return jsonify({'ok': False, 'erro': 'Nome vazio'}), 400
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO tamanhos_estoque (nome) VALUES (%s) ON CONFLICT DO NOTHING", (nome,))
    conn.commit(); cur.close(); close_db(conn)
    return jsonify({'ok': True, 'nome': nome})


@login_required
def etiquetas():
    data = request.args.get('data', hoje_app().isoformat())
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT codigo,modelo,descricao,tamanho,valor_venda,quantidade FROM estoque WHERE DATE(criado_em)=%s AND ativo=TRUE ORDER BY id", (data,))
    itens = [dict(i) for i in cur.fetchall()]
    cur.close(); close_db(conn)
    return jsonify({'itens': itens, 'data': data})


@login_required
def etiqueta_busca():
    # v142: código completo agora (ex.: PL5, SL12) — não tem mais um prefixo único "P"
    # pra completar sozinho.
    cod = request.args.get('codigo', '').strip().upper()
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id,codigo,modelo,descricao,tamanho,valor_venda,quantidade FROM estoque WHERE codigo=%s AND ativo=TRUE", (cod,))
    item = cur.fetchone(); cur.close(); close_db(conn)
    if item: return jsonify({'ok': True, 'item': dict(item)})
    return jsonify({'ok': False})


@login_required
def ficha_estoque(eid):
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM estoque WHERE id=%s", (eid,))
    row = cur.fetchone(); cur.close(); close_db(conn)
    if not row:
        flash('Produto nao encontrado.', 'erro'); return redirect(url_for('estoque'))
    item = dict(row)
    item['dias_estoque'] = (hoje_app() - item['criado_em'].date()).days
    item['saidas'] = (item['estoque_inicial'] or 0) - item['quantidade']
    ctx = get_ctx(); ctx['item'] = item
    return render_template('ficha_estoque.html', **ctx)


@login_required
def editar_estoque(eid):
    conn = get_db(); cur = conn.cursor()
    if request.method == 'POST':
        qtd = int(request.form.get('quantidade', 1) or 1)
        campos = (request.form.get('modelo', '').strip(),
                  request.form.get('descricao', '').strip() or None,
                  request.form.get('tamanho', '').strip(), qtd,
                  parse_brl(request.form.get('custo_unitario', '0')),
                  parse_brl(request.form.get('markup', '0')),
                  parse_brl(request.form.get('valor_venda', '0')),
                  parse_brl(request.form.get('margem_lucro', '0')))
        # Só mexe na foto se o usuário trocou/removeu (senão preserva a existente).
        if request.form.get('foto_alterada') == '1':
            foto = request.form.get('foto', '').strip() or None
            cur.execute("""UPDATE estoque SET modelo=%s,descricao=%s,tamanho=%s,quantidade=%s,
                custo_unitario=%s,markup=%s,valor_venda=%s,margem_lucro=%s,foto=%s WHERE id=%s""",
                campos + (foto, eid))
        else:
            cur.execute("""UPDATE estoque SET modelo=%s,descricao=%s,tamanho=%s,quantidade=%s,
                custo_unitario=%s,markup=%s,valor_venda=%s,margem_lucro=%s WHERE id=%s""",
                campos + (eid,))
        conn.commit(); cur.close(); close_db(conn)
        flash('Produto atualizado!', 'ok')
        return redirect(url_for('ficha_estoque', eid=eid))
    cur.execute("SELECT * FROM estoque WHERE id=%s", (eid,))
    item = cur.fetchone()
    if not item:
        cur.close(); close_db(conn)
        flash('Produto nao encontrado.', 'erro'); return redirect(url_for('estoque'))
    cur.execute("SELECT nome FROM modelos_estoque ORDER BY nome")
    modelos = [r['nome'] for r in cur.fetchall()]
    cur.execute("SELECT nome FROM tamanhos_estoque ORDER BY id")
    tamanhos = [r['nome'] for r in cur.fetchall()]
    cur.close(); close_db(conn)
    ctx = get_ctx(); ctx.update(item=item, modelos=modelos, tamanhos=tamanhos)
    return render_template('editar_estoque.html', **ctx)


@login_required
def aplicar_promocao():
    """Aplica ou remove o % de desconto promocional em vários produtos de uma vez.
    O preço original (valor_venda) NÃO é alterado — a promoção é só uma camada."""
    ids_raw = request.form.get('ids', '')
    ids = [int(x) for x in ids_raw.split(',') if x.strip().isdigit()]
    acao = request.form.get('acao', 'aplicar')
    if not ids:
        flash('Selecione ao menos um produto.', 'erro')
        return redirect(url_for('estoque'))
    # Modo de desconto: por VALOR (R$ fixo, convertido em % por produto) ou por %.
    valor_desc = parse_brl(request.form.get('valor_desconto', '0')) if acao != 'remover' else 0.0
    pct = 0.0
    if acao != 'remover' and valor_desc <= 0:
        pct = parse_brl(request.form.get('percentual', '0'))
        if pct <= 0 or pct >= 100:
            flash('Informe um percentual (1 a 99) ou um valor de desconto.', 'erro')
            return redirect(url_for('estoque'))
    conn = get_db(); cur = conn.cursor()
    try:
        if acao != 'remover' and valor_desc > 0:
            # R$ fixo: cada produto recebe o % equivalente ao seu próprio preço.
            cur.execute("SELECT id, valor_venda FROM estoque WHERE id = ANY(%s)", (ids,))
            for r in cur.fetchall():
                preco = float(r['valor_venda'] or 0)
                p = round(valor_desc / preco * 100, 2) if preco > 0 else 0.0
                p = max(0.0, min(99.0, p))
                cur.execute("UPDATE estoque SET desconto_promo=%s WHERE id=%s", (p, r['id']))
            flash(f'Desconto de R$ {valor_desc:.2f} aplicado em {len(ids)} produto(s).', 'ok')
        else:
            cur.execute("UPDATE estoque SET desconto_promo=%s WHERE id = ANY(%s)", (pct, ids))
            if acao == 'remover':
                flash(f'Promoção removida de {len(ids)} produto(s).', 'ok')
            else:
                flash(f'Desconto de {pct:.0f}% aplicado em {len(ids)} produto(s).', 'ok')
        conn.commit()
    except Exception as e:
        conn.rollback(); flash(str(e), 'erro')
    finally:
        cur.close(); close_db(conn)
    return redirect(url_for('estoque'))


@login_required
def excluir_estoque(eid):
    if not pode_excluir():
        flash('Apenas o Administrador N1 pode excluir dados.', 'erro'); return redirect(url_for('estoque'))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT codigo,modelo,descricao,tamanho,quantidade FROM estoque WHERE id=%s", (eid,))
    _old = cur.fetchone()
    cur.execute("DELETE FROM estoque WHERE id=%s", (eid,))
    audit_log(cur, 'EXCLUIR_PRODUTO', 'estoque', eid, dict(_old) if _old else None)
    conn.commit(); cur.close(); close_db(conn)
    flash('Produto excluido.', 'ok')
    return redirect(url_for('estoque'))


@login_required
def exportar_estoque():
    """v142: gera um .xlsx com os produtos do período filtrado (mesmo filtro da tela de
    Estoque — data de lançamento) para download. Uma linha por produto, com saldo,
    entradas/saídas, custo/venda e a promoção vigente."""
    conn = get_db(); cur = conn.cursor()
    hoje = hoje_app()
    data_inicio = request.args.get('data_inicio', hoje.strftime('%Y-01-01'))
    data_fim = request.args.get('data_fim', hoje.strftime('%Y-%m-%d'))
    try: date.fromisoformat(data_inicio)
    except ValueError: data_inicio = hoje.strftime('%Y-01-01')
    try: date.fromisoformat(data_fim)
    except ValueError: data_fim = hoje.strftime('%Y-%m-%d')
    cur.execute("SELECT * FROM estoque WHERE ativo=TRUE AND DATE(criado_em) BETWEEN %s AND %s ORDER BY criado_em", (data_inicio, data_fim))
    itens = [dict(i) for i in cur.fetchall()]
    cur.execute("SELECT estoque_id, COALESCE(SUM(quantidade),0) as total FROM estoque_entradas GROUP BY estoque_id")
    entradas_map = {r['estoque_id']: int(r['total']) for r in cur.fetchall()}
    cur.close(); close_db(conn)

    wb = _montar_wb_estoque(itens, entradas_map, hoje)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_arquivo = f"estoque_{data_inicio}_a_{data_fim}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome_arquivo,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@login_required
def exportar_fotos_estoque():
    """v143: extrai as FOTOS (guardadas em base64 no banco) dos produtos cujo
    código bate com a coluna 'Código' de um .xlsx (o mesmo formato do botão
    'Exportar dados', já tratado pelo usuário). Gera um .zip com um arquivo de
    imagem por código — usado pra levar as fotos separadamente antes de migrar
    os DADOS de uma loja pro ERP unificado (a importação de estoque não carrega foto)."""
    if request.method == 'GET':
        return render_template('exportar_fotos_estoque.html', **get_ctx())
    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename:
        flash('Selecione o arquivo .xlsx exportado.', 'erro')
        return redirect(url_for('exportar_fotos_estoque'))
    try:
        wb = load_workbook(io.BytesIO(arquivo.read()), data_only=True)
        ws = wb['Estoque'] if 'Estoque' in wb.sheetnames else wb.active
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            flash('Planilha vazia.', 'erro')
            return redirect(url_for('exportar_fotos_estoque'))
        cabecalho = [str(h).strip() if h is not None else '' for h in linhas[0]]
        col = {nome: i for i, nome in enumerate(cabecalho)}
        if 'Código' not in col:
            flash('A planilha precisa ter a coluna "Código" (use o arquivo gerado pelo botão Exportar dados).', 'erro')
            return redirect(url_for('exportar_fotos_estoque'))
        codigos = []
        for row in linhas[1:]:
            idx = col['Código']
            v = row[idx] if idx < len(row) else None
            if v is not None and str(v).strip():
                codigos.append(str(v).strip())
    except Exception as e:
        flash(f'Erro ao ler a planilha: {e}', 'erro')
        return redirect(url_for('exportar_fotos_estoque'))
    if not codigos:
        flash('Nenhum código encontrado na planilha.', 'erro')
        return redirect(url_for('exportar_fotos_estoque'))

    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT codigo, foto FROM estoque WHERE codigo = ANY(%s)", (codigos,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); close_db(conn)
    com_foto = [r for r in rows if r.get('foto')]
    codigos_encontrados = {r['codigo'] for r in rows}
    sem_foto = [c for c in codigos if c not in {r['codigo'] for r in com_foto}]
    nao_encontrados = [c for c in codigos if c not in codigos_encontrados]

    if not com_foto:
        flash('Nenhum dos códigos da planilha tem foto cadastrada no sistema.', 'erro')
        return redirect(url_for('exportar_fotos_estoque'))

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        usados = set()
        for r in com_foto:
            decoded = _foto_bytes_ext(r['foto'])
            if not decoded:
                continue
            dados, ext = decoded
            nome = r['codigo']
            sufixo = 1
            nome_final = nome
            while nome_final in usados:
                sufixo += 1
                nome_final = f"{nome}_{sufixo}"
            usados.add(nome_final)
            zf.writestr(f"{nome_final}.{ext}", dados)
        avisos = []
        if sem_foto:
            avisos.append('Códigos SEM foto cadastrada:\n' + '\n'.join(sem_foto))
        if nao_encontrados:
            avisos.append('Códigos não encontrados no sistema:\n' + '\n'.join(nao_encontrados))
        if avisos:
            zf.writestr('_avisos.txt', '\n\n'.join(avisos))
    buf.seek(0)
    if sem_foto or nao_encontrados:
        flash(f"{len(com_foto)} foto(s) exportada(s). {len(sem_foto)} código(s) sem foto e "
              f"{len(nao_encontrados)} não encontrado(s) — detalhes no arquivo _avisos.txt do zip.", 'ok')
    nome_arquivo = f"fotos_estoque_{hoje_app().strftime('%Y%m%d')}.zip"
    return send_file(buf, as_attachment=True, download_name=nome_arquivo, mimetype='application/zip')


@login_required
def importar_estoque():
    """v143: importa um .xlsx (o mesmo formato do botão 'Exportar dados', já tratado
    pelo usuário) + opcionalmente um .zip de fotos (do botão 'Exportar fotos') pro
    ERP unificado. Protegida pelos mesmos 2 cadeados de Clientes: env var
    MODO_MIGRACAO=true (só existe no serviço unificado) + admin N1.

    O admin escolhe o TIPO (Acessórios/Plus Size/Slim) na tela — define o prefixo do
    código NOVO (sequência própria, continua do maior já existente nesse tipo) e o
    valor gravado em tipo_produto. O código ANTIGO da planilha fica em origem_codigo,
    só como rastreio. Idempotente: reimportar o mesmo arquivo não duplica (índice
    único tipo_produto+origem_codigo). Se o .zip de fotos trouxer um arquivo cujo
    nome (sem extensão/pasta) bate com o código antigo, ela é decodificada e
    associada automaticamente ao produto novo — sem precisar subir foto por foto."""
    if not migracao_habilitada():
        flash('Importação não habilitada nesta instância.', 'erro'); return redirect(url_for('estoque'))
    if not pode_excluir():
        flash('Apenas o Administrador N1 pode importar dados.', 'erro'); return redirect(url_for('estoque'))
    if request.method == 'GET':
        ctx = get_ctx()
        ctx.update(tipos_produto=list(PREFIXOS_TIPO.keys()))
        return render_template('importar_estoque.html', **ctx)

    conn = get_db(); cur = conn.cursor()
    try:
        tipo_produto = request.form.get('tipo_produto', '').strip()
        prefixo = PREFIXOS_TIPO.get(tipo_produto)
        if not prefixo:
            flash('Selecione o tipo do produto (Acessórios, Plus Size ou Slim).', 'erro')
            return redirect(url_for('importar_estoque'))

        arquivo = request.files.get('arquivo')
        if not arquivo or not arquivo.filename:
            flash('Selecione o arquivo .xlsx exportado.', 'erro')
            return redirect(url_for('importar_estoque'))

        # .zip de fotos é opcional — sem ele, os produtos entram sem foto (dá pra
        # subir depois, um a um, na edição do produto).
        fotos_map = {}
        fotos_zip = request.files.get('fotos')
        if fotos_zip and fotos_zip.filename:
            with zipfile.ZipFile(io.BytesIO(fotos_zip.read())) as zf:
                for nome in zf.namelist():
                    base = nome.rsplit('/', 1)[-1]
                    if '.' not in base or base.startswith('_'):
                        continue
                    codigo_base, ext = base.rsplit('.', 1)
                    fotos_map[codigo_base.strip().upper()] = (zf.read(nome), ext.lower())

        wb = load_workbook(io.BytesIO(arquivo.read()), data_only=True)
        ws = wb['Estoque'] if 'Estoque' in wb.sheetnames else wb.active
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            flash('Planilha vazia.', 'erro')
            return redirect(url_for('importar_estoque'))
        cabecalho = [str(h).strip() if h is not None else '' for h in linhas[0]]
        col = {nome: i for i, nome in enumerate(cabecalho)}
        if 'Código' not in col:
            flash('A planilha precisa ter a coluna "Código" (use o arquivo gerado pelo botão Exportar dados).', 'erro')
            return redirect(url_for('importar_estoque'))

        def val(row, chave):
            i = col.get(chave)
            if i is None or i >= len(row):
                return None
            v = row[i]
            return v.strip() if isinstance(v, str) else v

        proximo_codigo = _proximo_codigo(cur, prefixo)
        importados = pulados = ignorados = com_foto = 0
        for row in linhas[1:]:
            origem_codigo = val(row, 'Código')
            if not origem_codigo or not str(origem_codigo).strip():
                ignorados += 1
                continue
            origem_codigo = str(origem_codigo).strip()
            saldo_raw = val(row, 'Saldo atual')
            try:
                saldo = int(saldo_raw) if saldo_raw is not None else 0
            except (ValueError, TypeError):
                saldo = 0
            custo = parse_brl(val(row, 'Custo unitário (R$)'))
            markup = parse_brl(val(row, 'Markup (%)'))
            venda = parse_brl(val(row, 'Valor de venda (R$)'))
            margem = parse_brl(val(row, 'Margem de lucro (%)'))
            criado_em = _parse_data_planilha(val(row, 'Data lançamento'))

            foto = None
            achou_foto = fotos_map.get(origem_codigo.upper())
            if achou_foto:
                dados, ext = achou_foto
                candidato = f"data:image/{ext};base64,{base64.b64encode(dados).decode('ascii')}"
                if len(candidato) <= 1_500_000:
                    foto = candidato

            novo_codigo = f"{prefixo}{proximo_codigo}"
            proximo_codigo += 1
            cur.execute("""INSERT INTO estoque
                (codigo,tipo_produto,modelo,descricao,tamanho,quantidade,estoque_inicial,
                 custo_unitario,markup,valor_venda,margem_lucro,foto,origem_codigo,criado_em)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,COALESCE(%s,CURRENT_TIMESTAMP))
                ON CONFLICT (tipo_produto, origem_codigo) WHERE origem_codigo IS NOT NULL DO NOTHING
                RETURNING id""",
                (novo_codigo, tipo_produto, val(row, 'Modelo'), val(row, 'Descrição'),
                 val(row, 'Tamanho'), saldo, saldo, custo, markup, venda, margem,
                 foto, origem_codigo, criado_em))
            if cur.fetchone():
                importados += 1
                if foto:
                    com_foto += 1
            else:
                pulados += 1
        conn.commit()
        audit_log(cur, 'IMPORTAR_ESTOQUE', 'estoque', None,
                  {'tipo_produto': tipo_produto, 'importados': importados, 'pulados': pulados})
        conn.commit()
        flash(f"Importação concluída: {importados} produto(s) novo(s) importado(s) "
              f"({com_foto} com foto), {pulados} já existia(m) (pulado(s) automaticamente), "
              f"{ignorados} linha(s) vazia(s) ignorada(s).", 'ok')
    except Exception as e:
        conn.rollback(); flash(f'Erro na importação: {e}', 'erro')
    finally:
        cur.close(); close_db(conn)
    return redirect(url_for('importar_estoque'))


def register(app):
    app.add_url_rule('/estoque', 'estoque', estoque)
    app.add_url_rule('/estoque/novo', 'novo_estoque', novo_estoque, methods=['POST'])
    app.add_url_rule('/estoque/<int:eid>/nova-entrada', 'nova_entrada_estoque', nova_entrada_estoque, methods=['POST'])
    app.add_url_rule('/estoque/modelo/novo', 'novo_modelo', novo_modelo, methods=['POST'])
    app.add_url_rule('/estoque/tamanho/novo', 'novo_tamanho', novo_tamanho, methods=['POST'])
    app.add_url_rule('/estoque/etiquetas', 'etiquetas', etiquetas)
    app.add_url_rule('/estoque/etiqueta-busca', 'etiqueta_busca', etiqueta_busca)
    app.add_url_rule('/estoque/promocao', 'aplicar_promocao', aplicar_promocao, methods=['POST'])
    app.add_url_rule('/estoque/exportar', 'exportar_estoque', exportar_estoque)
    app.add_url_rule('/estoque/exportar-fotos', 'exportar_fotos_estoque', exportar_fotos_estoque, methods=['GET', 'POST'])
    app.add_url_rule('/estoque/importar', 'importar_estoque', importar_estoque, methods=['GET', 'POST'])
    app.add_url_rule('/estoque/<int:eid>', 'ficha_estoque', ficha_estoque)
    app.add_url_rule('/estoque/<int:eid>/editar', 'editar_estoque', editar_estoque, methods=['GET', 'POST'])
    app.add_url_rule('/estoque/<int:eid>/excluir', 'excluir_estoque', excluir_estoque, methods=['POST'])
